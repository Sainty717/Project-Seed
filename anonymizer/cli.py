"""
Command-line interface for the anonymization framework
"""

import click
from pathlib import Path
from typing import List, Optional
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn
from rich.prompt import Confirm, Prompt
from datetime import datetime

from .core.detector import DataTypeDetector, DataType
from .core.vault import MappingVault
from .core.transformers import FormatPreservingTransformer
from .utils.csv_processor import CSVProcessor
from .utils.excel_processor import ExcelProcessor
from .utils.validators import ValidationReport
from .config.profiles import AnonymizationProfile, AnonymizationMode, get_default_profiles


console = Console()


@click.group()
@click.version_option(version="1.0.0")
def cli():
    """Format-Preserving Data Anonymization Framework"""
    pass


@cli.command()
@click.option('--input', '-i', required=True, multiple=True, help='Input file(s) - CSV or Excel (.xlsx, .xls, .xlsm, .xlsb, .ods)')
@click.option('--output', '-o', required=True, help='Output directory')
@click.option('--profile', '-p', default='default', help='Anonymization profile')
@click.option('--columns', '-c', multiple=True, help='Columns to anonymize (all if not specified)')
@click.option('--interactive', '-I', is_flag=True, help='Interactive column selection')
@click.option('--seed', '-s', help='Deterministic seed for anonymization')
@click.option('--mode', '-m', type=click.Choice(['fake', 'fpe', 'hmac', 'hybrid']), help='Anonymization mode')
@click.option('--vault', '-v', help='Path to existing mapping vault (creates new if not specified)')
@click.option('--vault-password', help='Password for mapping vault encryption')
@click.option('--preview/--no-preview', default=True, help='Show preview before processing')
@click.option('--preserve-domain', is_flag=True, help='Preserve email domains')
@click.option('--no-vault', is_flag=True, help='Do not store mappings (fully synthetic)')
# Excel-specific options
@click.option('--sheet', multiple=True, help='Sheet name(s) to process (all sheets if not specified, Excel only)')
@click.option('--merge-sheets', is_flag=True, help='Merge multiple sheets into one sheet (Excel only, default: preserve sheet structure)')
@click.option('--separate-sheets', is_flag=True, help='Write each sheet to a separate file (Excel only, default: preserve sheet structure)')
@click.option('--header-row', type=int, help='Row index (0-based) to use as header (auto-detect if not specified, Excel only)')
@click.option('--skip-rows', type=int, default=0, help='Number of rows to skip before reading (Excel only)')
@click.option('--output-format', type=click.Choice(['excel', 'csv']), default='excel', help='Output format for Excel files (excel or csv)')
def anonymize(
    input: tuple,
    output: str,
    profile: str,
    columns: tuple,
    interactive: bool,
    seed: Optional[str],
    mode: Optional[str],
    vault: Optional[str],
    vault_password: Optional[str],
    preview: bool,
    preserve_domain: bool,
    no_vault: bool,
    sheet: tuple,
    merge_sheets: bool,
    separate_sheets: bool,
    header_row: Optional[int],
    skip_rows: int,
    output_format: str
):
    """Anonymize CSV or Excel file(s) while preserving format"""
    
    input_files = list(input)
    output_dir = Path(output)
    
    # Create output directory structure
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    session_dir = output_dir / timestamp
    anonymized_dir = session_dir / "anonymized_files"
    original_dir = session_dir / "original_files"
    anonymized_dir.mkdir(parents=True, exist_ok=True)
    original_dir.mkdir(parents=True, exist_ok=True)
    
    console.print(f"\n[bold green]Starting anonymization session[/bold green]")
    console.print(f"Output directory: {session_dir}\n")
    
    # Load or create profile
    profiles = get_default_profiles()
    if profile in profiles:
        anonymization_profile = profiles[profile]
    else:
        console.print(f"[yellow]Profile '{profile}' not found, using default[/yellow]")
        anonymization_profile = profiles["default"]
    
    # Override profile settings from CLI
    if seed:
        anonymization_profile.seed = seed
    if mode:
        mode_map = {
            'fake': AnonymizationMode.FORMAT_PRESERVING_FAKE,
            'fpe': AnonymizationMode.FPE,
            'hmac': AnonymizationMode.SEEDED_HMAC,
            'hybrid': AnonymizationMode.HYBRID
        }
        anonymization_profile.mode = mode_map[mode]
    if preserve_domain:
        anonymization_profile.preserve_domain = True
    if no_vault:
        anonymization_profile.fully_synthetic = True
    
    # Initialize vault
    vault_obj = None
    if not anonymization_profile.fully_synthetic:
        if vault:
            # Use existing vault
            vault_path = Path(vault)
            if not vault_path.exists():
                console.print(f"[yellow]Warning: Vault file not found at {vault_path}, creating new vault[/yellow]")
                vault_path.parent.mkdir(parents=True, exist_ok=True)
            
            vault_obj = MappingVault(str(vault_path), vault_password)
            console.print(f"[green]✓[/green] Using existing vault: {vault_path}")
        else:
            # Create new vault in session directory
            vault_path = session_dir / "mapping_vault.sqlite"
            vault_obj = MappingVault(str(vault_path), vault_password)
            console.print(f"[green]✓[/green] Mapping vault initialized: {vault_path}")
    
    # Create transformer
    transformer = anonymization_profile.create_transformer(vault=vault_obj)
    
    console.print(f"[green]✓[/green] Transformer created: {anonymization_profile.mode.value}")
    
    # Detect file types and initialize appropriate processors
    excel_files = []
    csv_files = []
    
    for input_file in input_files:
        input_path = Path(input_file)
        if ExcelProcessor.is_excel_file(input_path):
            excel_files.append(input_file)
        else:
            csv_files.append(input_file)
    
    # Initialize processors
    csv_processor = CSVProcessor(transformer=transformer) if csv_files else None
    excel_processor = ExcelProcessor(transformer=transformer) if excel_files else None
    
    if excel_files:
        console.print(f"[green]✓[/green] Excel processor initialized ({len(excel_files)} Excel file(s))")
    if csv_files:
        console.print(f"[green]✓[/green] CSV processor initialized ({len(csv_files)} CSV file(s))")
    
    # Interactive column selection
    columns_to_anonymize = list(columns) if columns else None
    excel_columns_by_sheet = {}  # Dict to store column selections per sheet for Excel files
    
    if interactive and input_files:
        console.print("\n[bold cyan]Interactive Column Selection[/bold cyan]")
        console.print("Analyzing file schema...\n")
        
        try:
            # Detect schema from first file
            first_file = input_files[0]
            first_path = Path(first_file)
            
            # Use appropriate processor based on file type
            if ExcelProcessor.is_excel_file(first_path):
                if not excel_processor:
                    console.print("[red]Error: Excel processor not initialized[/red]")
                    return
                
                # For Excel, iterate through each sheet
                sheets = excel_processor.list_sheets(first_file, include_hidden=False)
                
                if not sheet:
                    # Auto-select all visible sheets for interactive mode
                    selected_sheets = [s['name'] for s in sheets if s['visible']]
                    console.print(f"[bold]Found {len(selected_sheets)} sheet(s) in {first_path.name}[/bold]\n")
                else:
                    selected_sheets = list(sheet)
                
                # Go through each sheet and let user select columns
                for sheet_name in selected_sheets:
                    console.print(f"\n[bold cyan]Sheet: {sheet_name}[/bold cyan]")
                    console.print("-" * 60)
                    
                    # Extract schema for this sheet
                    schema = excel_processor.extract_schema(
                        first_file,
                        sheet_name=sheet_name,
                        header_row=header_row,
                        skip_rows=skip_rows
                    )
                    
                    # Display schema table
                    schema_table = Table(title=f"Columns in '{sheet_name}'", show_header=True, header_style="bold magenta")
                    schema_table.add_column("#", style="dim", width=4)
                    schema_table.add_column("Column Name", style="cyan")
                    schema_table.add_column("Detected Type", style="green")
                    schema_table.add_column("Confidence", style="yellow")
                    
                    column_list = list(schema.keys())
                    for idx, (col_name, (data_type, confidence)) in enumerate(schema.items(), 1):
                        schema_table.add_row(
                            str(idx),
                            col_name,
                            data_type.value,
                            f"{confidence:.1%}"
                        )
                    
                    console.print(schema_table)
                    console.print()
                    
                    # Interactive selection for this sheet
                    if columns:
                        # If columns were provided via CLI, use those for all sheets
                        sheet_columns = list(columns)
                        console.print(f"[green]Using columns from command line: {', '.join(sheet_columns)}[/green]")
                    else:
                        # Let user select columns for this sheet
                        console.print(f"[bold]Select columns to anonymize in '{sheet_name}':[/bold]")
                        console.print("[dim]Enter column numbers separated by commas (e.g., 1,2,3) or 'all' for all columns[/dim]")
                        console.print("[dim]Press Enter with no input to anonymize all columns[/dim]\n")
                        
                        selection = Prompt.ask(f"Column selection for '{sheet_name}'", default="all")
                        
                        if selection.lower() == 'all' or not selection.strip():
                            sheet_columns = column_list
                            console.print(f"[green]Selected all columns: {', '.join(sheet_columns)}[/green]")
                        else:
                            try:
                                # Parse comma-separated numbers
                                indices = [int(x.strip()) for x in selection.split(',')]
                                selected_columns = [column_list[i-1] for i in indices if 1 <= i <= len(column_list)]
                                
                                if selected_columns:
                                    sheet_columns = selected_columns
                                    console.print(f"[green]Selected columns: {', '.join(sheet_columns)}[/green]")
                                else:
                                    console.print("[yellow]No valid columns selected, anonymizing all columns[/yellow]")
                                    sheet_columns = column_list
                            except (ValueError, IndexError):
                                console.print("[red]Invalid selection format. Using all columns.[/red]")
                                sheet_columns = column_list
                    
                    excel_columns_by_sheet[sheet_name] = sheet_columns
                    console.print()
                
                # Update sheet tuple with selected sheets
                sheet = tuple(selected_sheets)
                console.print(f"[green]✓[/green] Completed column selection for {len(selected_sheets)} sheet(s)\n")
                
            else:
                # CSV file - standard interactive mode
                if not csv_processor:
                    console.print("[red]Error: CSV processor not initialized[/red]")
                    return
                
                schema = csv_processor.extract_schema(first_file)
                
                # Display schema table
                schema_table = Table(title="Detected Columns", show_header=True, header_style="bold magenta")
                schema_table.add_column("#", style="dim", width=4)
                schema_table.add_column("Column Name", style="cyan")
                schema_table.add_column("Detected Type", style="green")
                schema_table.add_column("Confidence", style="yellow")
                
                column_list = list(schema.keys())
                for idx, (col_name, (data_type, confidence)) in enumerate(schema.items(), 1):
                    schema_table.add_row(
                        str(idx),
                        col_name,
                        data_type.value,
                        f"{confidence:.1%}"
                    )
                
                console.print(schema_table)
                console.print()
                
                # Interactive selection
                if columns:
                    # If columns were provided via CLI, use those
                    console.print(f"[green]Using columns from command line: {', '.join(columns)}[/green]")
                    columns_to_anonymize = list(columns)
                else:
                    # Let user select columns
                    console.print("[bold]Select columns to anonymize:[/bold]")
                    console.print("[dim]Enter column numbers separated by commas (e.g., 1,2,3) or 'all' for all columns[/dim]")
                    console.print("[dim]Press Enter with no input to anonymize all columns[/dim]\n")
                    
                    selection = Prompt.ask("Column selection", default="all")
                    
                    if selection.lower() == 'all' or not selection.strip():
                        columns_to_anonymize = column_list
                        console.print(f"[green]Selected all columns: {', '.join(columns_to_anonymize)}[/green]")
                    else:
                        try:
                            # Parse comma-separated numbers
                            indices = [int(x.strip()) for x in selection.split(',')]
                            selected_columns = [column_list[i-1] for i in indices if 1 <= i <= len(column_list)]
                            
                            if selected_columns:
                                columns_to_anonymize = selected_columns
                                console.print(f"[green]Selected columns: {', '.join(columns_to_anonymize)}[/green]")
                            else:
                                console.print("[yellow]No valid columns selected, anonymizing all columns[/yellow]")
                                columns_to_anonymize = column_list
                        except (ValueError, IndexError):
                            console.print("[red]Invalid selection format. Using all columns.[/red]")
                            columns_to_anonymize = column_list
                    
                    console.print()
        
        except Exception as e:
            console.print(f"[red]Error during interactive selection: {e}[/red]")
            console.print("[yellow]Falling back to anonymizing all columns[/yellow]")
            columns_to_anonymize = None
            excel_columns_by_sheet = {}
    
    # Show preview if requested
    if preview and input_files:
        console.print("\n[bold]Preview Mode[/bold]")
        preview_file = input_files[0]
        preview_path = Path(preview_file)
        
        try:
            if ExcelProcessor.is_excel_file(preview_path):
                if not excel_processor:
                    console.print("[red]Error: Excel processor not initialized[/red]")
                    return
                
                # For Excel with multiple sheets, show preview for first sheet
                sheet_name = sheet[0] if sheet else None
                
                # Use per-sheet columns if available
                preview_columns = None
                if excel_columns_by_sheet and sheet_name and sheet_name in excel_columns_by_sheet:
                    preview_columns = excel_columns_by_sheet[sheet_name]
                else:
                    preview_columns = columns_to_anonymize
                
                preview_df = excel_processor.preview_transformation(
                    preview_file,
                    sheet_name=sheet_name,
                    columns_to_anonymize=preview_columns,
                    num_samples=5,
                    header_row=header_row,
                    skip_rows=skip_rows
                )
                
                if sheet_name:
                    console.print(f"[dim]Preview for sheet: {sheet_name}[/dim]\n")
            else:
                if not csv_processor:
                    console.print("[red]Error: CSV processor not initialized[/red]")
                    return
                preview_df = csv_processor.preview_transformation(
                    preview_file,
                    columns_to_anonymize,
                    num_samples=5
                )
            
            # Display preview table
            table = Table(title="Transformation Preview", show_header=True, header_style="bold magenta")
            
            # Add columns
            for col in preview_df.columns:
                table.add_column(col, overflow="fold")
            
            # Add rows
            for _, row in preview_df.head(5).iterrows():
                table.add_row(*[str(val) for val in row])
            
            console.print(table)
            
            if not Confirm.ask("\nProceed with anonymization?"):
                console.print("[yellow]Cancelled by user[/yellow]")
                return
        
        except Exception as e:
            console.print(f"[red]Error generating preview: {e}[/red]")
            if not Confirm.ask("Continue anyway?"):
                return
    
    # Initialize validation report
    validation_report = ValidationReport(str(session_dir))
    
    # Process files
    console.print(f"\n[bold]Processing {len(input_files)} file(s)...[/bold]\n")
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        console=console
    ) as progress:
        task = progress.add_task("Processing files...", total=len(input_files))
        
        for input_file in input_files:
            input_path = Path(input_file)
            
            if not input_path.exists():
                console.print(f"[red]Error: File not found: {input_file}[/red]")
                validation_report.add_error(f"File not found: {input_file}")
                continue
            
            try:
                # Copy original to original_files directory
                import shutil
                shutil.copy2(input_path, original_dir / input_path.name)
                
                # Process file based on type
                if ExcelProcessor.is_excel_file(input_path):
                    if not excel_processor:
                        console.print(f"[red]Error: Excel processor not initialized for {input_path.name}[/red]")
                        continue
                    
                    # Process Excel file
                    # If no sheets specified and not in interactive mode, get all sheets
                    if not sheet and not interactive:
                        all_sheets = excel_processor.list_sheets(str(input_path), include_hidden=False)
                        sheet_names = [s['name'] for s in all_sheets if s['visible']]
                    else:
                        sheet_names = list(sheet) if sheet else None
                    
                    # Determine output behavior
                    if merge_sheets:
                        # Merge all sheets into one sheet (explicit merge)
                        results = excel_processor.process_multiple_sheets(
                            str(input_path),
                            str(anonymized_dir),
                            sheet_names=sheet_names,
                            merge_sheets=True,
                            columns_to_anonymize=columns_to_anonymize,
                            header_row=header_row,
                            skip_rows=skip_rows,
                            show_progress=True,
                            output_format=output_format
                        )
                        
                        for result in results:
                            validation_report.add_file_result(
                                f"{input_path.stem}_merged",
                                result["columns_anonymized"],
                                result["rows_processed"]
                            )
                        console.print(f"[green]✓[/green] Processed: {input_path.name} - Merged {len(sheet_names) if sheet_names else 'all'} sheet(s)")
                    
                    elif separate_sheets and sheet_names and len(sheet_names) > 1:
                        # Explicit request for separate files
                        results = excel_processor.process_multiple_sheets(
                            str(input_path),
                            str(anonymized_dir),
                            sheet_names=sheet_names,
                            merge_sheets=False,
                            columns_to_anonymize=columns_to_anonymize,
                            header_row=header_row,
                            skip_rows=skip_rows,
                            show_progress=True,
                            output_format=output_format
                        )
                        
                        for result in results:
                            validation_report.add_file_result(
                                f"{input_path.stem}_{result['sheet_name']}",
                                result["columns_anonymized"],
                                result["rows_processed"]
                            )
                            console.print(f"[green]✓[/green] Processed: {input_path.name} - Sheet: {result['sheet_name']}")
                    
                    elif sheet_names and len(sheet_names) > 1:
                        # Multiple sheets - default: preserve structure (one Excel file with multiple sheets)
                        output_file = anonymized_dir / f"{input_path.stem}.xlsx"
                        
                        # Use per-sheet column selections if available from interactive mode
                        columns_dict = excel_columns_by_sheet if excel_columns_by_sheet else None
                        if columns_dict is None and columns_to_anonymize:
                            # If same columns for all sheets, create dict
                            columns_dict = {name: columns_to_anonymize for name in sheet_names}
                        
                        result = excel_processor.process_multiple_sheets_to_one_file(
                            str(input_path),
                            str(output_file),
                            sheet_names=sheet_names,
                            columns_to_anonymize=columns_dict,
                            header_row=header_row,
                            skip_rows=skip_rows,
                            show_progress=True
                        )
                        
                        # Add results for each sheet
                        for sheet_name, sheet_result in result["results_by_sheet"].items():
                            validation_report.add_file_result(
                                f"{input_path.stem}::{sheet_name}",
                                sheet_result["columns_anonymized"],
                                sheet_result["rows"]
                            )
                        
                        console.print(f"[green]✓[/green] Processed: {input_path.name} - {len(sheet_names)} sheet(s) preserved in one Excel file")
                    
                    else:
                        # Single sheet
                        sheet_name = sheet_names[0] if sheet_names else None
                        output_file = anonymized_dir / f"{input_path.stem}.{'xlsx' if output_format == 'excel' else 'csv'}"
                        
                        result = excel_processor.process_sheet(
                            str(input_path),
                            str(output_file),
                            sheet_name=sheet_name,
                            columns_to_anonymize=columns_to_anonymize,
                            header_row=header_row,
                            skip_rows=skip_rows,
                            show_progress=True,
                            output_format=output_format
                        )
                        
                        validation_report.add_file_result(
                            input_path.name,
                            result["columns_anonymized"],
                            result["rows_processed"]
                        )
                        console.print(f"[green]✓[/green] Processed: {input_path.name}")
                else:
                    # Process CSV file
                    if not csv_processor:
                        console.print(f"[red]Error: CSV processor not initialized for {input_path.name}[/red]")
                        continue
                    
                    output_file = anonymized_dir / input_path.name
                    result = csv_processor.process_file(
                        str(input_path),
                        str(output_file),
                        columns_to_anonymize=columns_to_anonymize,
                        show_progress=True
                    )
                    
                    validation_report.add_file_result(
                        input_path.name,
                        result["columns_anonymized"],
                        result["rows_processed"]
                    )
                    console.print(f"[green]✓[/green] Processed: {input_path.name}")
                
                progress.update(task, advance=1)
            
            except Exception as e:
                error_msg = f"Error processing {input_file}: {str(e)}"
                console.print(f"[red]✗ {error_msg}[/red]")
                validation_report.add_error(error_msg)
                progress.update(task, advance=1)
    
    # Save format rules
    format_rules_path = session_dir / "format_rules_used.json"
    import json
    with open(format_rules_path, 'w') as f:
        json.dump({
            "profile": anonymization_profile.name,
            "mode": anonymization_profile.mode.value,
            "seed": anonymization_profile.seed,
            "columns_anonymized": columns_to_anonymize or "all",
            "timestamp": timestamp
        }, f, indent=2)
    
    console.print(f"[green]✓[/green] Format rules saved: {format_rules_path}")
    
    # Export decryption key if vault exists (only for new vaults, not existing ones)
    if vault_obj and not anonymization_profile.fully_synthetic and not vault:
        key_path = session_dir / "decryption_key.json"
        vault_obj.export_key(str(key_path))
        console.print(f"[yellow]⚠[/yellow] Decryption key saved: {key_path}")
        console.print("[yellow]⚠[/yellow] Keep this key secure for reversibility!")
    elif vault_obj and vault:
        # If using existing vault, remind user about key file location
        console.print(f"[dim]Using existing vault. If you need the decryption key, use the key file from when the vault was created.[/dim]")
    
    # Generate validation report
    report_path = validation_report.generate_report()
    console.print(f"[green]✓[/green] Validation report: {report_path}")
    
    # Show summary
    console.print("\n[bold green]Anonymization Complete![/bold green]\n")
    
    summary_table = Table(title="Summary", show_header=True, header_style="bold cyan")
    summary_table.add_column("Metric", style="cyan")
    summary_table.add_column("Value", style="green")
    
    summary_table.add_row("Files processed", str(len(validation_report.report_data["files_processed"])))
    summary_table.add_row("Total rows", str(sum(f["rows"] for f in validation_report.report_data["files_processed"])))
    summary_table.add_row("Columns anonymized", str(len(validation_report.report_data["columns_anonymized"])))
    summary_table.add_row("Output directory", str(session_dir))
    
    console.print(summary_table)


@cli.command()
@click.option('--file', '-f', required=True, help='File to analyze (CSV or Excel)')
@click.option('--sample', '-s', default=100, help='Number of rows to sample')
@click.option('--sheet', help='Sheet name to analyze (Excel only, uses first sheet if not specified)')
@click.option('--header-row', type=int, help='Row index (0-based) to use as header (Excel only, auto-detect if not specified)')
@click.option('--skip-rows', type=int, default=0, help='Number of rows to skip before reading (Excel only)')
def analyze(file: str, sample: int, sheet: Optional[str], header_row: Optional[int], skip_rows: int):
    """Analyze CSV or Excel file and detect data types"""
    
    file_path = Path(file)
    if not file_path.exists():
        console.print(f"[red]Error: File not found: {file}[/red]")
        return
    
    console.print(f"\n[bold]Analyzing: {file}[/bold]\n")
    
    detector = DataTypeDetector()
    
    if ExcelProcessor.is_excel_file(file_path):
        processor = ExcelProcessor(transformer=None, detector=detector)
        schema = processor.extract_schema(
            file,
            sheet_name=sheet,
            sample_rows=sample,
            header_row=header_row,
            skip_rows=skip_rows
        )
        
        if sheet:
            console.print(f"[dim]Sheet: {sheet}[/dim]\n")
    else:
        processor = CSVProcessor(transformer=None, detector=detector)
        schema = processor.extract_schema(file, sample_rows=sample)
    
    # Display results
    table = Table(title="Schema Detection Results", show_header=True, header_style="bold magenta")
    table.add_column("Column", style="cyan")
    table.add_column("Detected Type", style="green")
    table.add_column("Confidence", style="yellow")
    
    for column, (data_type, confidence) in schema.items():
        confidence_str = f"{confidence:.1%}"
        table.add_row(column, data_type.value, confidence_str)
    
    console.print(table)


@cli.command()
@click.option('--vault', '-v', required=True, help='Path to mapping vault')
@click.option('--password', '-p', help='Vault password')
@click.option('--original', '-o', required=True, help='Original value')
@click.option('--column', '-c', required=True, help='Column name')
@click.option('--seed', '-s', help='Seed used for anonymization')
def reverse(vault: str, password: Optional[str], original: str, column: str, seed: Optional[str]):
    """Reverse lookup: get anonymized value from original"""
    
    vault_obj = MappingVault(vault, password)
    anonymized = vault_obj.get_mapping(original, column, seed)
    
    if anonymized:
        console.print(f"\n[green]Original:[/green] {original}")
        console.print(f"[green]Anonymized:[/green] {anonymized}\n")
    else:
        console.print(f"[red]No mapping found for: {original} in column: {column}[/red]")


@cli.command()
@click.option('--input', '-i', required=True, help='Anonymized file to decrypt (CSV or Excel)')
@click.option('--output', '-o', required=True, help='Output decrypted file (CSV or Excel)')
@click.option('--vault', '-v', required=True, help='Path to mapping vault')
@click.option('--password', '-p', help='Vault password')
@click.option('--key-file', '-k', help='Path to decryption key JSON file')
@click.option('--columns', '-c', multiple=True, help='Columns to decrypt (all anonymized columns if not specified)')
@click.option('--seed', '-s', help='Seed used for anonymization')
@click.option('--sheet', multiple=True, help='Sheet name(s) to decrypt (Excel only, all sheets if not specified)')
def decrypt(
    input: str,
    output: str,
    vault: str,
    password: Optional[str],
    key_file: Optional[str],
    columns: tuple,
    seed: Optional[str],
    sheet: tuple
):
    """Decrypt anonymized CSV or Excel file back to original values"""
    
    import pandas as pd
    from tqdm import tqdm
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    input_path = Path(input)
    output_path = Path(output)
    
    if not input_path.exists():
        console.print(f"[red]Error: File not found: {input}[/red]")
        return
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    console.print(f"\n[bold green]Decrypting file: {input}[/bold green]")
    console.print(f"Output: {output}\n")
    
    # Load vault
    try:
        vault_obj = MappingVault(vault, password)
        
        # Load key if provided (overrides password)
        if key_file:
            key_path = Path(key_file)
            if not key_path.exists():
                console.print(f"[red]Error: Key file not found: {key_file}[/red]")
                return
            vault_obj.load_key(key_file)
            console.print(f"[green]✓[/green] Loaded decryption key from: {key_file}")
        elif password:
            console.print(f"[green]✓[/green] Using vault password")
        else:
            console.print("[yellow]Warning: No password or key file provided. Trying to open vault...[/yellow]")
    
    except Exception as e:
        console.print(f"[red]Error loading vault: {e}[/red]")
        console.print("[yellow]Make sure you have the correct password or key file[/yellow]")
        return
    
    # Check if input is Excel file
    is_excel = ExcelProcessor.is_excel_file(input_path)
    
    if is_excel:
        # Process Excel file
        excel_processor = ExcelProcessor(transformer=None)
        
        # Get list of sheets
        all_sheets = excel_processor.list_sheets(str(input_path), include_hidden=False)
        
        # Determine which sheets to process
        if sheet:
            sheet_names = [s['name'] for s in all_sheets if s['name'] in sheet and s['visible']]
            if not sheet_names:
                console.print(f"[yellow]Warning: No valid sheets found. Available sheets: {[s['name'] for s in all_sheets if s['visible']]}[/yellow]")
                return
        else:
            # Process all visible sheets
            sheet_names = [s['name'] for s in all_sheets if s['visible']]
        
        console.print(f"[green]✓[/green] Found {len(sheet_names)} sheet(s) to decrypt")
        
        # Get columns with mappings from vault
        columns_with_mappings = set()
        try:
            stats = vault_obj.get_statistics()
            columns_with_mappings = set(stats.get("column_counts", {}).keys())
        except Exception as e:
            console.print(f"[yellow]Could not query vault for column mappings: {e}[/yellow]")
        
        # Process each sheet
        total_decrypted_count = 0
        total_not_found_count = 0
        total_failed_count = 0
        total_rows = 0
        all_columns_decrypted = set()
        results_by_sheet = {}
        
        # Create new workbook for output
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for sheet_name in tqdm(sheet_names, desc="Decrypting sheets", disable=False):
            console.print(f"\n[bold]Processing sheet: {sheet_name}[/bold]")
            
            try:
                # Read sheet
                df = excel_processor.read_excel_sheet(
                    str(input_path),
                    sheet_name=sheet_name
                )
                
                total_rows += len(df)
                
                # Determine columns to decrypt for this sheet
                if columns:
                    # User specified columns - decrypt only those
                    columns_to_decrypt = [col for col in columns if col in df.columns]
                else:
                    # Auto-detect: only decrypt columns that have mappings in the vault
                    if columns_with_mappings:
                        columns_to_decrypt = [col for col in df.columns if col in columns_with_mappings]
                    else:
                        # Try all columns
                        columns_to_decrypt = list(df.columns)
                
                # Filter to only columns that exist
                columns_to_decrypt = [col for col in columns_to_decrypt if col in df.columns]
                
                if not columns_to_decrypt:
                    console.print(f"[dim]No columns to decrypt in sheet '{sheet_name}' (preserving as-is)[/dim]")
                    # Still add sheet to output
                    ws = wb.create_sheet(title=sheet_name)
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    results_by_sheet[sheet_name] = {
                        "rows": len(df),
                        "columns_decrypted": [],
                        "values_decrypted": 0
                    }
                    continue
                
                all_columns_decrypted.update(columns_to_decrypt)
                console.print(f"[dim]Decrypting {len(columns_to_decrypt)} column(s) in '{sheet_name}'...[/dim]")
                
                # Decrypt each column
                decrypted_count = 0
                not_found_count = 0
                failed_count = 0
                
                for column in columns_to_decrypt:
                    try:
                        original_count = len(df[column].dropna())
                        decrypted_in_col = 0
                        
                        # Decrypt values
                        def decrypt_value(x):
                            nonlocal decrypted_in_col
                            if pd.notna(x) and str(x).strip():
                                original = vault_obj.reverse_lookup(str(x), column, seed)
                                if original is not None:
                                    decrypted_in_col += 1
                                    return original
                                # If not found in vault, keep the anonymized value
                            return x
                        
                        df[column] = df[column].apply(decrypt_value)
                        
                        decrypted_count += decrypted_in_col
                        not_found = original_count - decrypted_in_col
                        not_found_count += not_found
                        
                        if decrypted_in_col > 0:
                            console.print(f"[green]  ✓[/green] {column}: {decrypted_in_col}/{original_count} values decrypted")
                        elif original_count > 0:
                            console.print(f"[yellow]  ⚠[/yellow] {column}: No mappings found in vault (kept as-is)")
                    
                    except Exception as e:
                        console.print(f"[red]  ✗[/red] Error decrypting {column}: {e}")
                        failed_count += 1
                
                total_decrypted_count += decrypted_count
                total_not_found_count += not_found_count
                total_failed_count += failed_count
                
                # Add sheet to workbook
                ws = wb.create_sheet(title=sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                results_by_sheet[sheet_name] = {
                    "rows": len(df),
                    "columns_decrypted": columns_to_decrypt,
                    "values_decrypted": decrypted_count
                }
                
            except Exception as e:
                console.print(f"[red]Error processing sheet '{sheet_name}': {e}[/red]")
                total_failed_count += 1
                continue
        
        # Save Excel file
        try:
            wb.save(str(output_path))
            console.print(f"\n[green]✓[/green] Decrypted Excel file saved to: {output_path}")
        except Exception as e:
            console.print(f"[red]Error saving Excel file: {e}[/red]")
            return
        
        # Summary
        console.print("\n[bold green]Decryption Complete![/bold green]\n")
        
        summary_table = Table(title="Summary", show_header=True, header_style="bold cyan")
        summary_table.add_column("Metric", style="cyan")
        summary_table.add_column("Value", style="green")
        
        summary_table.add_row("Sheets processed", str(len(sheet_names)))
        summary_table.add_row("Total rows processed", str(total_rows))
        summary_table.add_row("Columns decrypted", str(len(all_columns_decrypted)))
        summary_table.add_row("Values decrypted", str(total_decrypted_count))
        if total_not_found_count > 0:
            summary_table.add_row("Values not found in vault", str(total_not_found_count), style="dim")
        if total_failed_count > 0:
            summary_table.add_row("Failed columns", str(total_failed_count), style="yellow")
        
        console.print(summary_table)
        
    else:
        # Process CSV file (original logic)
        try:
            df = pd.read_csv(input_path)
            console.print(f"[green]✓[/green] Loaded {len(df)} rows from CSV")
        except Exception as e:
            console.print(f"[red]Error reading CSV: {e}[/red]")
            return
        
        # Determine columns to decrypt
        if columns:
            # User specified columns - decrypt only those
            columns_to_decrypt = [col for col in columns if col in df.columns]
        else:
            # Auto-detect: only decrypt columns that have mappings in the vault
            try:
                stats = vault_obj.get_statistics()
                columns_with_mappings = set(stats.get("column_counts", {}).keys())
                # Only decrypt columns that exist in both CSV and vault
                columns_to_decrypt = [col for col in df.columns if col in columns_with_mappings]
                
                if columns_to_decrypt:
                    console.print(f"[dim]Auto-detected {len(columns_to_decrypt)} column(s) with mappings in vault[/dim]")
                else:
                    console.print("[yellow]No columns with mappings found in vault. Nothing to decrypt.[/yellow]")
                    console.print("[dim]All columns will be preserved as-is.[/dim]")
            except Exception as e:
                console.print(f"[yellow]Could not query vault for column mappings: {e}[/yellow]")
                console.print("[yellow]Will attempt to decrypt all columns...[/yellow]")
                columns_to_decrypt = list(df.columns)
        
        # Filter to only columns that exist
        columns_to_decrypt = [col for col in columns_to_decrypt if col in df.columns]
        
        if not columns_to_decrypt:
            console.print("[yellow]No valid columns to decrypt. Saving file with all original values preserved.[/yellow]")
            # Still save the file even if nothing to decrypt
            df.to_csv(output_path, index=False)
            console.print(f"[green]✓[/green] File saved to: {output_path} (no decryption needed)")
            return
        
        console.print(f"\n[bold]Decrypting {len(columns_to_decrypt)} column(s)...[/bold]\n")
        
        # Decrypt each column
        decrypted_count = 0
        failed_count = 0
        not_found_count = 0
        
        for column in tqdm(columns_to_decrypt, desc="Decrypting columns", disable=False):
            try:
                original_count = len(df[column].dropna())
                decrypted_in_col = 0
                
                # Decrypt values
                def decrypt_value(x):
                    nonlocal decrypted_in_col
                    if pd.notna(x) and str(x).strip():
                        original = vault_obj.reverse_lookup(str(x), column, seed)
                        if original is not None:
                            decrypted_in_col += 1
                            return original
                        # If not found in vault, keep the anonymized value
                    return x
                
                df[column] = df[column].apply(decrypt_value)
                
                decrypted_count += decrypted_in_col
                not_found = original_count - decrypted_in_col
                
                if decrypted_in_col > 0:
                    console.print(f"[green]✓[/green] {column}: {decrypted_in_col}/{original_count} values decrypted")
                    if not_found > 0:
                        console.print(f"[dim]  ({not_found} values not found in vault, kept as-is)[/dim]")
                else:
                    console.print(f"[yellow]⚠[/yellow] {column}: No mappings found in vault (kept as-is)")
                    not_found_count += not_found
            
            except Exception as e:
                console.print(f"[red]✗[/red] Error decrypting {column}: {e}")
                failed_count += 1
        
        # Save decrypted file
        try:
            df.to_csv(output_path, index=False)
            console.print(f"\n[green]✓[/green] Decrypted file saved to: {output_path}")
        except Exception as e:
            console.print(f"[red]Error saving file: {e}[/red]")
            return
        
        # Summary
        console.print("\n[bold green]Decryption Complete![/bold green]\n")
        
        summary_table = Table(title="Summary", show_header=True, header_style="bold cyan")
        summary_table.add_column("Metric", style="cyan")
        summary_table.add_column("Value", style="green")
        
        summary_table.add_row("Rows processed", str(len(df)))
        summary_table.add_row("Total columns", str(len(df.columns)))
        summary_table.add_row("Columns decrypted", str(len(columns_to_decrypt)))
        summary_table.add_row("Values decrypted", str(decrypted_count))
        if not_found_count > 0:
            summary_table.add_row("Values not found in vault", str(not_found_count), style="dim")
        if failed_count > 0:
            summary_table.add_row("Failed columns", str(failed_count), style="yellow")
        
        console.print(summary_table)


@cli.command()
def profiles():
    """List available anonymization profiles"""
    
    profiles_dict = get_default_profiles()
    
    table = Table(title="Available Profiles", show_header=True, header_style="bold magenta")
    table.add_column("Name", style="cyan")
    table.add_column("Mode", style="green")
    table.add_column("Reversible", style="yellow")
    table.add_column("Description", style="white")
    
    for name, profile in profiles_dict.items():
        reversible = "Yes" if profile.mode != AnonymizationMode.SEEDED_HMAC and not profile.fully_synthetic else "No"
        description = {
            "default": "Balanced hybrid approach",
            "gdpr_compliant": "FPE with reversible mappings",
            "test_data": "Synthetic data generation",
            "fast_hash": "Fast non-reversible hashing",
            "referential_integrity": "Maintains cross-dataset consistency"
        }.get(name, "")
        
        table.add_row(name, profile.mode.value, reversible, description)
    
    console.print(table)


if __name__ == '__main__':
    cli()

