"""
Excel processing utilities with support for multiple formats and sheet handling
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Union
from tqdm import tqdm
import warnings
import re

from ..core.detector import DataTypeDetector, DataType
from ..core.transformers import FormatPreservingTransformer

# Suppress openpyxl warnings about merged cells
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class ExcelProcessor:
    """Process Excel files with schema detection and anonymization"""
    
    # Supported Excel file extensions
    EXCEL_EXTENSIONS = {'.xlsx', '.xls', '.xlsm', '.xlsb', '.ods'}
    
    def __init__(
        self,
        transformer: FormatPreservingTransformer,
        detector: Optional[DataTypeDetector] = None,
        chunk_size: int = 10000,
        use_read_only: bool = True
    ):
        """
        Initialize Excel processor
        
        Args:
            transformer: Transformer to use for anonymization
            detector: Optional data type detector (creates new if None)
            chunk_size: Size of chunks for processing large files
            use_read_only: Use read-only mode for openpyxl (faster, less memory)
        """
        self.transformer = transformer
        self.detector = detector or DataTypeDetector()
        self.chunk_size = chunk_size
        self.use_read_only = use_read_only
    
    @staticmethod
    def is_excel_file(file_path: Union[str, Path]) -> bool:
        """Check if file is an Excel file based on extension"""
        path = Path(file_path)
        return path.suffix.lower() in ExcelProcessor.EXCEL_EXTENSIONS
    
    def list_sheets(
        self,
        file_path: str,
        include_hidden: bool = False
    ) -> List[Dict[str, Any]]:
        """
        List all sheets in an Excel file
        
        Args:
            file_path: Path to Excel file
            include_hidden: Whether to include hidden sheets
            
        Returns:
            List of dicts with sheet info: {'name': str, 'visible': bool, 'index': int}
        """
        file_path_obj = Path(file_path)
        ext = file_path_obj.suffix.lower()
        
        sheets = []
        
        try:
            if ext == '.xlsb':
                # Binary Excel format
                try:
                    import pyxlsb
                    with pyxlsb.open_workbook(file_path) as wb:
                        for idx, sheet_name in enumerate(wb.sheets):
                            sheets.append({
                                'name': sheet_name,
                                'visible': True,  # pyxlsb doesn't expose visibility
                                'index': idx
                            })
                except ImportError:
                    raise ImportError("pyxlsb is required for .xlsb files. Install with: pip install pyxlsb")
            elif ext == '.ods':
                # OpenDocument Spreadsheet - use pandas with odf engine
                try:
                    xl_file = pd.ExcelFile(file_path, engine='odf')
                    for idx, sheet_name in enumerate(xl_file.sheet_names):
                        sheets.append({
                            'name': sheet_name,
                            'visible': True,
                            'index': idx
                        })
                except Exception as e:
                    raise ValueError(f"Could not read ODS file. Install odfpy: pip install odfpy. Error: {e}")
            else:
                # .xlsx, .xls, .xlsm - use openpyxl
                wb = openpyxl.load_workbook(
                    file_path,
                    read_only=self.use_read_only,
                    data_only=True  # Get calculated values, not formulas
                )
                
                for idx, sheet_name in enumerate(wb.sheetnames):
                    sheet = wb[sheet_name]
                    is_visible = sheet.sheet_state == 'visible'
                    
                    if include_hidden or is_visible:
                        sheets.append({
                            'name': sheet_name,
                            'visible': is_visible,
                            'index': idx
                        })
                
                wb.close()
        
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {e}")
        
        return sheets
    
    def detect_header_row(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        skip_rows: int = 0,
        max_rows_to_check: int = 20
    ) -> int:
        """
        Automatically detect header row in Excel sheet
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (uses first sheet if None)
            skip_rows: Number of rows to skip before checking
            max_rows_to_check: Maximum rows to check for header
        
        Returns:
            Row index (0-based) where header is detected
        """
        try:
            # Read first few rows
            df_sample = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                nrows=max_rows_to_check,
                header=None,
                skiprows=skip_rows
            )
            
            # Find first row with mostly non-null values
            for idx in range(len(df_sample)):
                row = df_sample.iloc[idx]
                non_null_count = row.notna().sum()
                total_count = len(row)
                
                # If row has >50% non-null values, likely a header
                if non_null_count > total_count * 0.5:
                    return skip_rows + idx
            
            # Default to first row after skip_rows
            return skip_rows
        
        except Exception as e:
            # Fallback to skip_rows
            return skip_rows
    
    def read_excel_sheet(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        header_row: Optional[int] = None,
        skip_rows: int = 0,
        nrows: Optional[int] = None,
        use_read_only: Optional[bool] = None,
        engine: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Read Excel sheet with proper handling of headers, merged cells, etc.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (uses first sheet if None)
            header_row: Row index (0-based) to use as header (auto-detect if None)
            skip_rows: Number of rows to skip before reading
            nrows: Number of rows to read (None for all)
            use_read_only: Override read-only mode
            engine: Pandas engine to use ('openpyxl', 'xlrd', 'odf', 'pyxlsb')
        
        Returns:
            DataFrame with data
        """
        file_path_obj = Path(file_path)
        ext = file_path_obj.suffix.lower()
        
        # Determine engine if not specified
        if engine is None:
            if ext == '.xls':
                engine = 'xlrd'
            elif ext == '.xlsb':
                engine = 'pyxlsb'
            elif ext == '.ods':
                engine = 'odf'
            else:
                engine = 'openpyxl'
        
        # Auto-detect header if not specified
        if header_row is None:
            header_row = self.detect_header_row(file_path, sheet_name, skip_rows)
        
        # Calculate skiprows and header position
        if header_row is not None and header_row > skip_rows:
            # Skip rows before header, use header row as header
            rows_to_skip = list(range(skip_rows, header_row))
            header_pos = 0  # After skipping, header is at position 0
        elif skip_rows > 0:
            # Skip rows, but header_row is at or before skip_rows, so no header
            rows_to_skip = list(range(skip_rows))
            header_pos = None  # No header row
        else:
            rows_to_skip = []
            header_pos = header_row if header_row is not None else 0
        
        read_kwargs = {
            'sheet_name': sheet_name or 0,
            'engine': engine,
            'nrows': nrows
        }
        
        if rows_to_skip:
            read_kwargs['skiprows'] = rows_to_skip
        
        if header_pos is not None:
            read_kwargs['header'] = header_pos
        else:
            read_kwargs['header'] = None
        
        try:
            df = pd.read_excel(file_path, **read_kwargs)
            
            # Clean up: remove fully empty rows and columns
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            # Clean column names (remove extra whitespace, handle merged cell artifacts)
            df.columns = [str(col).strip() if pd.notna(col) else f'Unnamed_{i}' 
                         for i, col in enumerate(df.columns)]
            
            # Remove duplicate column names
            df.columns = self._fix_duplicate_columns(df.columns)
            
            return df
        
        except Exception as e:
            raise ValueError(f"Error reading Excel sheet '{sheet_name}': {e}")
    
    def _fix_duplicate_columns(self, columns: pd.Index) -> List[str]:
        """Fix duplicate column names by appending numbers"""
        seen = {}
        result = []
        
        for col in columns:
            if col in seen:
                seen[col] += 1
                result.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                result.append(col)
        
        return result
    
    def extract_schema(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        sample_rows: int = 100,
        header_row: Optional[int] = None,
        skip_rows: int = 0
    ) -> Dict[str, Tuple[DataType, float]]:
        """
        Extract schema from Excel file
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (uses first sheet if None)
            sample_rows: Number of rows to sample for detection
            header_row: Row index to use as header
            skip_rows: Number of rows to skip before reading
        
        Returns:
            Dictionary mapping column names to (type, confidence) tuples
        """
        df_sample = self.read_excel_sheet(
            file_path,
            sheet_name=sheet_name,
            header_row=header_row,
            skip_rows=skip_rows,
            nrows=sample_rows
        )
        
        schema = self.detector.detect_schema(df_sample, sample_size=min(sample_rows, len(df_sample)))
        return schema
    
    def process_sheet(
        self,
        file_path: str,
        output_path: str,
        sheet_name: Optional[str] = None,
        columns_to_anonymize: Optional[List[str]] = None,
        schema_override: Optional[Dict[str, DataType]] = None,
        header_row: Optional[int] = None,
        skip_rows: int = 0,
        show_progress: bool = True,
        output_format: str = 'excel'  # 'excel' or 'csv'
    ) -> Dict[str, Any]:
        """
        Process a single Excel sheet and anonymize specified columns
        
        Args:
            file_path: Path to input Excel file
            output_path: Path to output file
            sheet_name: Sheet name to process (uses first sheet if None)
            columns_to_anonymize: List of column names to anonymize (all if None)
            schema_override: Optional manual type overrides per column
            header_row: Row index to use as header (auto-detect if None)
            skip_rows: Number of rows to skip before reading
            show_progress: Whether to show progress bar
            output_format: Output format ('excel' or 'csv')
        
        Returns:
            Dictionary with processing statistics
        """
        input_path_obj = Path(file_path)
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        # Detect schema
        schema = self.extract_schema(
            file_path,
            sheet_name=sheet_name,
            header_row=header_row,
            skip_rows=skip_rows
        )
        
        # Apply overrides
        if schema_override:
            for col, data_type in schema_override.items():
                if col in schema:
                    schema[col] = (data_type, 1.0)
        
        # Determine columns to process
        if columns_to_anonymize is None:
            columns_to_anonymize = list(schema.keys())
        
        # Filter to only columns that exist
        columns_to_anonymize = [col for col in columns_to_anonymize if col in schema]
        
        # Read full sheet (or in chunks for large files)
        file_size = input_path_obj.stat().st_size
        file_size_mb = file_size / (1024 * 1024)
        
        # For Excel files, we'll process in memory but show progress
        # Note: pandas read_excel doesn't support true chunked reading like CSV
        # For very large files, consider converting to CSV first or using streaming libraries
        return self._process_sheet_full(
            file_path,
            output_path,
            sheet_name,
            columns_to_anonymize,
            schema,
            header_row,
            skip_rows,
            show_progress,
            output_format
        )
    
    def _process_sheet_full(
        self,
        file_path: str,
        output_path: str,
        sheet_name: Optional[str],
        columns_to_anonymize: List[str],
        schema: Dict[str, Tuple[DataType, float]],
        header_row: Optional[int],
        skip_rows: int,
        show_progress: bool,
        output_format: str
    ) -> Dict[str, Any]:
        """Process entire sheet at once"""
        # Read full sheet
        df = self.read_excel_sheet(
            file_path,
            sheet_name=sheet_name,
            header_row=header_row,
            skip_rows=skip_rows
        )
        
        total_rows = len(df)
        
        # Anonymize columns
        with tqdm(total=total_rows, desc="Processing", disable=not show_progress) as pbar:
            for column in columns_to_anonymize:
                if column in df.columns:
                    data_type, _ = schema[column]
                    
                    # Apply transformation
                    df[column] = df[column].apply(
                        lambda x: self.transformer.transform(
                            x,
                            data_type,
                            column
                        ) if pd.notna(x) and str(x).strip() else x
                    )
            
            pbar.update(total_rows)
        
        # Write output
        if output_format == 'excel':
            df.to_excel(output_path, index=False, engine='openpyxl')
        else:
            df.to_csv(output_path, index=False)
        
        return {
            "input_file": file_path,
            "output_file": output_path,
            "sheet_name": sheet_name or "Sheet1",
            "rows_processed": total_rows,
            "columns_anonymized": columns_to_anonymize,
            "chunks_processed": 1
        }
    
    def _process_sheet_chunked(
        self,
        file_path: str,
        output_path: str,
        sheet_name: Optional[str],
        columns_to_anonymize: List[str],
        schema: Dict[str, Tuple[DataType, float]],
        header_row: Optional[int],
        skip_rows: int,
        show_progress: bool,
        output_format: str
    ) -> Dict[str, Any]:
        """
        Process sheet in chunks for very large files.
        Note: This is a workaround since pandas read_excel doesn't support true chunking.
        For best performance with very large Excel files, consider converting to CSV first.
        """
        # For now, fall back to full processing with a warning
        import warnings
        warnings.warn(
            "Chunked processing for Excel files is limited. "
            "Processing entire file in memory. For very large files, consider converting to CSV first.",
            UserWarning
        )
        return self._process_sheet_full(
            file_path,
            output_path,
            sheet_name,
            columns_to_anonymize,
            schema,
            header_row,
            skip_rows,
            show_progress,
            output_format
        )
    
    
    def process_multiple_sheets(
        self,
        file_path: str,
        output_dir: str,
        sheet_names: Optional[List[str]] = None,
        merge_sheets: bool = False,
        columns_to_anonymize: Optional[List[str]] = None,
        header_row: Optional[int] = None,
        skip_rows: int = 0,
        show_progress: bool = True,
        output_format: str = 'excel'
    ) -> List[Dict[str, Any]]:
        """
        Process multiple sheets from an Excel file
        
        Args:
            file_path: Path to input Excel file
            output_dir: Directory for output files
            sheet_names: List of sheet names to process (all if None)
            merge_sheets: If True, merge all sheets into one output file
            columns_to_anonymize: Columns to anonymize
            header_row: Row index to use as header
            skip_rows: Number of rows to skip
            show_progress: Whether to show progress
            output_format: Output format ('excel' or 'csv')
        
        Returns:
            List of processing statistics for each sheet
        """
        output_dir_path = Path(output_dir)
        output_dir_path.mkdir(parents=True, exist_ok=True)
        
        # Get list of sheets
        all_sheets = self.list_sheets(file_path, include_hidden=False)
        
        if sheet_names is None:
            sheet_names = [s['name'] for s in all_sheets]
        
        # Filter to only visible sheets that exist
        available_sheets = {s['name']: s for s in all_sheets}
        sheet_names = [name for name in sheet_names if name in available_sheets]
        
        if not sheet_names:
            raise ValueError("No valid sheets found to process")
        
        results = []
        
        if merge_sheets:
            # Merge all sheets into one output
            input_path_obj = Path(file_path)
            output_file = output_dir_path / f"{input_path_obj.stem}_merged.{'xlsx' if output_format == 'excel' else 'csv'}"
            
            all_dataframes = []
            
            for sheet_name in tqdm(sheet_names, desc="Reading sheets", disable=not show_progress):
                df = self.read_excel_sheet(
                    file_path,
                    sheet_name=sheet_name,
                    header_row=header_row,
                    skip_rows=skip_rows
                )
                
                # Add sheet name as column if merging
                df.insert(0, '_Sheet', sheet_name)
                all_dataframes.append(df)
            
            # Combine all dataframes
            combined_df = pd.concat(all_dataframes, ignore_index=True)
            
            # Anonymize
            schema = self.detector.detect_schema(combined_df)
            if columns_to_anonymize is None:
                columns_to_anonymize = list(schema.keys())
            
            for column in columns_to_anonymize:
                if column in combined_df.columns:
                    data_type, _ = schema[column]
                    combined_df[column] = combined_df[column].apply(
                        lambda x: self.transformer.transform(
                            x,
                            data_type,
                            column
                        ) if pd.notna(x) and str(x).strip() else x
                    )
            
            # Write output
            if output_format == 'excel':
                combined_df.to_excel(output_file, index=False, engine='openpyxl')
            else:
                combined_df.to_csv(output_file, index=False)
            
            results.append({
                "input_file": file_path,
                "output_file": str(output_file),
                "sheet_name": "merged",
                "rows_processed": len(combined_df),
                "columns_anonymized": columns_to_anonymize,
                "sheets_merged": len(sheet_names)
            })
        
        else:
            # Process each sheet separately
            for sheet_name in tqdm(sheet_names, desc="Processing sheets", disable=not show_progress):
                input_path_obj = Path(file_path)
                output_file = output_dir_path / f"{input_path_obj.stem}_{sheet_name}.{'xlsx' if output_format == 'excel' else 'csv'}"
                
                result = self.process_sheet(
                    file_path,
                    str(output_file),
                    sheet_name=sheet_name,
                    columns_to_anonymize=columns_to_anonymize,
                    header_row=header_row,
                    skip_rows=skip_rows,
                    show_progress=False,
                    output_format=output_format
                )
                results.append(result)
        
        return results
    
    def preview_transformation(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        columns_to_anonymize: Optional[List[str]] = None,
        num_samples: int = 10,
        header_row: Optional[int] = None,
        skip_rows: int = 0
    ) -> pd.DataFrame:
        """
        Generate preview of transformations
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to preview
            columns_to_anonymize: Columns to preview
            num_samples: Number of sample rows to show
            header_row: Row index to use as header
            skip_rows: Number of rows to skip
        
        Returns:
            DataFrame with original and anonymized columns side by side
        """
        df = self.read_excel_sheet(
            file_path,
            sheet_name=sheet_name,
            header_row=header_row,
            skip_rows=skip_rows,
            nrows=num_samples
        )
        
        schema = self.extract_schema(
            file_path,
            sheet_name=sheet_name,
            header_row=header_row,
            skip_rows=skip_rows
        )
        
        if columns_to_anonymize is None:
            columns_to_anonymize = list(schema.keys())
        
        preview_data = {}
        
        for column in df.columns:
            preview_data[f"{column}_original"] = df[column]
            
            if column in columns_to_anonymize and column in schema:
                data_type, _ = schema[column]
                preview_data[f"{column}_anonymized"] = df[column].apply(
                    lambda x: self.transformer.transform(
                        x,
                        data_type,
                        column
                    ) if pd.notna(x) and str(x).strip() else x
                )
            else:
                preview_data[f"{column}_anonymized"] = df[column]
        
        return pd.DataFrame(preview_data)
    
    def process_multiple_sheets_to_one_file(
        self,
        file_path: str,
        output_path: str,
        sheet_names: Optional[List[str]] = None,
        columns_to_anonymize: Optional[Dict[str, List[str]]] = None,
        header_row: Optional[int] = None,
        skip_rows: int = 0,
        show_progress: bool = True
    ) -> Dict[str, Any]:
        """
        Process multiple sheets and write them to a single Excel file preserving sheet structure
        
        Args:
            file_path: Path to input Excel file
            output_path: Path to output Excel file (will contain multiple sheets)
            sheet_names: List of sheet names to process (all if None)
            columns_to_anonymize: Dict mapping sheet_name -> list of columns to anonymize
                                  If None, anonymize all columns for all sheets
            header_row: Row index to use as header
            skip_rows: Number of rows to skip
            show_progress: Whether to show progress
        
        Returns:
            Dictionary with processing statistics
        """
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        input_path_obj = Path(file_path)
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        # Get list of sheets
        all_sheets = self.list_sheets(file_path, include_hidden=False)
        
        if sheet_names is None:
            sheet_names = [s['name'] for s in all_sheets]
        
        # Filter to only visible sheets that exist
        available_sheets = {s['name']: s for s in all_sheets}
        sheet_names = [name for name in sheet_names if name in available_sheets]
        
        if not sheet_names:
            raise ValueError("No valid sheets found to process")
        
        # Create new workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        total_rows = 0
        all_columns_anonymized = set()
        results_by_sheet = {}
        
        # Process each sheet
        for sheet_name in tqdm(sheet_names, desc="Processing sheets", disable=not show_progress):
            # Read sheet
            df = self.read_excel_sheet(
                file_path,
                sheet_name=sheet_name,
                header_row=header_row,
                skip_rows=skip_rows
            )
            
            # Get columns to anonymize for this sheet
            if columns_to_anonymize and sheet_name in columns_to_anonymize:
                sheet_columns = columns_to_anonymize[sheet_name]
            elif columns_to_anonymize is None:
                # Anonymize all columns
                schema = self.extract_schema(
                    file_path,
                    sheet_name=sheet_name,
                    header_row=header_row,
                    skip_rows=skip_rows
                )
                sheet_columns = list(schema.keys())
            else:
                sheet_columns = []
            
            # Anonymize columns
            if sheet_columns:
                schema = self.extract_schema(
                    file_path,
                    sheet_name=sheet_name,
                    header_row=header_row,
                    skip_rows=skip_rows
                )
                
                for column in sheet_columns:
                    if column in df.columns:
                        data_type, _ = schema.get(column, (DataType.FREE_TEXT, 0.5))
                        df[column] = df[column].apply(
                            lambda x: self.transformer.transform(
                                x,
                                data_type,
                                column
                            ) if pd.notna(x) and str(x).strip() else x
                        )
                        all_columns_anonymized.add(column)
            
            # Create new sheet in workbook
            ws = wb.create_sheet(title=sheet_name)
            
            # Write dataframe to sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            total_rows += len(df)
            results_by_sheet[sheet_name] = {
                "rows": len(df),
                "columns_anonymized": sheet_columns
            }
        
        # Save workbook
        wb.save(output_path)
        
        return {
            "input_file": file_path,
            "output_file": output_path,
            "sheets_processed": len(sheet_names),
            "total_rows_processed": total_rows,
            "columns_anonymized": list(all_columns_anonymized),
            "results_by_sheet": results_by_sheet
        }

